#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import re
import openpyxl

wb = openpyxl.load_workbook('/Users/cherishxn/workspace/pycharm_projects/douyin-mcp-server/美食爆款钩子.xlsx')
ws = wb.active

links = []
for row in range(2, ws.max_row + 1):
    val = ws.cell(row=row, column=1).value
    if val and isinstance(val, str):
        # 提取 https://v.douyin.com/xxx/ 或 https://www.douyin.com/xxx 链接
        found = re.findall(r'https?://(?:v\.douyin\.com/[^\s]+|www\.douyin\.com/[^\s]+)', val)
        if found:
            links.append({"row": row, "link": found[0], "text": val[:60]})

print(f"共提取到 {len(links)} 条视频链接\n")
for item in links[:10]:
    print(f"  行{item['row']}: {item['link']}")
    print(f"         {item['text']}...")
    print()

if len(links) > 10:
    print(f"  ... 还有 {len(links)-10} 条")

# 保存到文件方便后续使用
with open('/Users/cherishxn/workspace/pycharm_projects/douyin-mcp-server/links.txt', 'w') as f:
    for item in links:
        f.write(f"{item['row']}\t{item['link']}\n")
print(f"\n链接已保存到 links.txt")
