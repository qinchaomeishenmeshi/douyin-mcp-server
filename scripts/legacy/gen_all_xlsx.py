#!/usr/bin/env python3
"""将全部 ASR 结果（100旧 + 136新）整理成 xlsx"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

RESULT_FILE = "/Users/cherishxn/workspace/pycharm_projects/douyin-mcp-server/美食爆款钩子_all_转文字结果.json"
LINKS_OLD = "/Users/cherishxn/workspace/pycharm_projects/douyin-mcp-server/video_links.json"
LINKS_NEW = "/Users/cherishxn/workspace/pycharm_projects/douyin-mcp-server/video_links_new.json"

with open(RESULT_FILE, 'r', encoding='utf-8') as f:
    data = json.load(f)

# 加载链接映射 (aweme_id -> link)
aweme_to_link = {}
for lf in [LINKS_OLD, LINKS_NEW]:
    if os.path.exists(lf):
        with open(lf, 'r', encoding='utf-8') as f:
            links = json.load(f)
            for item in links:
                aid = item.get('aweme_id', '')
                link = item.get('link', '')
                if aid and link:
                    aweme_to_link[aid] = link

wb = Workbook()
ws = wb.active
ws.title = "美食爆款钩子-全部转文字"

header_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='E74C3C')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_font = Font(name='Arial', size=10)
cell_align = Alignment(vertical='top', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

headers = ['序号', '视频名称', '时长(秒)', '语音转文字', '原始链接']
col_widths = [8, 45, 10, 80, 40]

for col, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws.column_dimensions[chr(64 + col)].width = width

for idx, r in enumerate(data, 2):
    aweme_id = r.get('aweme_id', '')
    link = aweme_to_link.get(aweme_id, r.get('original_link', ''))
    row_data = [
        r.get('index', ''),
        r.get('video_name', ''),
        r.get('duration', ''),
        r.get('text', ''),
        link,
    ]
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=idx, column=col, value=val)
        cell.font = cell_font
        cell.alignment = cell_align
        cell.border = thin_border
        if idx % 2 == 0:
            cell.fill = PatternFill('solid', fgColor='FFF5F5')

ws.freeze_panes = 'A2'

output_path = '/Users/cherishxn/workspace/pycharm_projects/douyin-mcp-server/美食爆款钩子_全部转文字结果.xlsx'
wb.save(output_path)
print(f"✅ XLSX 已保存: {output_path}")
print(f"共 {len(data)} 条记录")
