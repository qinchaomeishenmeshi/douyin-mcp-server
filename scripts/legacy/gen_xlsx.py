#!/usr/bin/env python3
"""将 ASR 结果整理成美观的 xlsx"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

with open('/Users/cherishxn/workspace/pycharm_projects/douyin-mcp-server/美食爆款钩子_转文字结果.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

wb = Workbook()
ws = wb.active
ws.title = "美食爆款钩子-转文字"

# 样式
header_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='E74C3C')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_font = Font(name='Arial', size=10)
cell_align = Alignment(vertical='top', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# 表头
headers = ['序号', '视频名称', '时长(秒)', '语音转文字', '原始链接']
col_widths = [8, 45, 10, 80, 40]

for col, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws.column_dimensions[chr(64 + col)].width = width

# 数据
for idx, r in enumerate(data, 2):
    row_data = [
        r.get('index', ''),
        r.get('video_name', ''),
        r.get('duration', ''),
        r.get('text', ''),
        r.get('original_link', ''),
    ]
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=idx, column=col, value=val)
        cell.font = cell_font
        cell.alignment = cell_align
        cell.border = thin_border
        # 交替行底色
        if idx % 2 == 0:
            cell.fill = PatternFill('solid', fgColor='FFF5F5')

# 冻结首行
ws.freeze_panes = 'A2'

output_path = '/Users/cherishxn/workspace/pycharm_projects/douyin-mcp-server/美食爆款钩子_转文字结果.xlsx'
wb.save(output_path)
print(f"✅ XLSX 已保存: {output_path}")
